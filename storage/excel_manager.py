from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
import os

FILE_PATH = "storage/leads.xlsx"
SHEET_NAME = "Leads"


def save_emails_to_excel(emails):
    if os.path.exists(FILE_PATH):
        wb = load_workbook(FILE_PATH)
        sheet = wb[SHEET_NAME]
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.title = SHEET_NAME

        headers = [
            # "Message ID",
            "From",
            "Subject",
            "Date",
            "Body",
            "Lead Category",
            "Priority",
            "Owner Action",
            "Auto Reply", 
            "Status"
        ]

        sheet.append(headers)

        # ✅ STICKY HEADER
        sheet.freeze_panes = "A2"

        # ✅ HEADER STYLE
        header_fill = PatternFill(
            start_color="BDD7EE",  # light blue
            end_color="BDD7EE",
            fill_type="solid"
        )

        header_font = Font(bold=True)

        header_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

    for email in emails:
        sheet.append([
            # email["message_id"],
            email["from"],
            email["subject"],
            email["date"],
            email["body"],
            email.get("category", ""),
            email.get("priority", ""),
            email.get("owner_action", ""),
            email.get("auto_reply", ""),
            "NEW"
        ])

        # ✅ WRAP BODY TEXT
        body_cell = sheet.cell(row=sheet.max_row, column=5)
        body_cell.alignment = Alignment(
            wrap_text=True,
            vertical="top"
        )

    # Column widths
    sheet.column_dimensions["A"].width = 40
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["C"].width = 50
    sheet.column_dimensions["D"].width = 25
    sheet.column_dimensions["E"].width = 80

    wb.save(FILE_PATH)
